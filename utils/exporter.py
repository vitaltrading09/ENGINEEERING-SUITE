"""
exporter.py
-----------
Reusable export utility for all calculators.
Generates professionally formatted PDF reports (ReportLab) and
structured Excel workbooks (openpyxl).
"""

from __future__ import annotations

import os
from datetime import datetime
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from calculators.base_calculator import BaseCalculator


class Exporter:
    """
    Static helper class.  Call as:
        Exporter.export_pdf(calculator_widget, "/path/to/file.pdf")
        Exporter.export_excel(calculator_widget, "/path/to/file.xlsx")
    """

    COMPANY = "Solareff Engineering"
    APP_NAME = "Engineering Calculator Suite"
    SANS_REF = "SANS 10142-1:2020"
    PRIMARY_COLOR = (31, 63, 96)      # deep navy  (R,G,B)
    ACCENT_COLOR  = (88, 166, 255)    # blue accent

    # ------------------------------------------------------------------
    # PDF Export
    # ------------------------------------------------------------------

    @staticmethod
    def export_pdf(calc: "BaseCalculator", filepath: str) -> None:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.lib import colors
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
            HRFlowable
        )
        from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT

        doc = SimpleDocTemplate(
            filepath,
            pagesize=A4,
            leftMargin=2.2 * cm,
            rightMargin=2.2 * cm,
            topMargin=2.2 * cm,
            bottomMargin=2.2 * cm,
        )

        # ---- Colours ----
        navy   = colors.Color(31/255, 63/255, 96/255)
        blue   = colors.Color(88/255, 166/255, 255/255)
        green  = colors.Color(48/255, 185/255, 80/255)
        red    = colors.Color(248/255, 81/255, 73/255)
        grey   = colors.Color(0.55, 0.55, 0.55)
        light  = colors.Color(0.94, 0.96, 0.98)
        white  = colors.white

        # ---- Styles ----
        styles = getSampleStyleSheet()
        style_h1 = ParagraphStyle("h1",
            fontSize=18, fontName="Helvetica-Bold",
            textColor=navy, spaceAfter=4)
        style_h2 = ParagraphStyle("h2",
            fontSize=12, fontName="Helvetica-Bold",
            textColor=navy, spaceBefore=12, spaceAfter=6)
        style_body = ParagraphStyle("body",
            fontSize=10, fontName="Helvetica",
            textColor=colors.black, spaceAfter=4, leading=14)
        style_small = ParagraphStyle("small",
            fontSize=8, fontName="Helvetica",
            textColor=grey, spaceAfter=2)
        style_footer = ParagraphStyle("footer",
            fontSize=8, fontName="Helvetica",
            textColor=grey, alignment=TA_CENTER)

        now = datetime.now().strftime("%d %B %Y  %H:%M")
        inputs  = calc.get_inputs()
        results = calc.get_results()

        story = []

        # ---- Header band ----
        header_data = [[
            Paragraph(f"<b>{Exporter.APP_NAME}</b>", style_h1),
            Paragraph(
                f"<font color='grey' size='9'>{Exporter.COMPANY}<br/>"
                f"Generated: {now}</font>",
                ParagraphStyle("hdr_r", fontSize=9, alignment=TA_RIGHT,
                               textColor=grey, fontName="Helvetica"))
        ]]
        header_table = Table(header_data, colWidths=["65%", "35%"])
        header_table.setStyle(TableStyle([
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
            ("LEFTPADDING",  (0,0), (-1,-1), 0),
            ("RIGHTPADDING", (0,0), (-1,-1), 0),
        ]))
        story.append(header_table)
        story.append(HRFlowable(width="100%", thickness=2, color=navy,
                                spaceAfter=6, spaceBefore=4))

        # Calculator name
        story.append(Paragraph(calc.calculator_name, style_h2))
        story.append(Paragraph(
            f"Reference Standard: <b>{calc.sans_reference}</b>", style_body))
        story.append(Spacer(1, 0.3*cm))

        # ---- Input Parameters Table ----
        story.append(Paragraph("Input Parameters", style_h2))
        story.append(HRFlowable(width="100%", thickness=0.5, color=blue,
                                spaceBefore=2, spaceAfter=6))

        in_rows = [["Parameter", "Value"]]
        for k, v in inputs.items():
            in_rows.append([k, str(v)])

        in_table = Table(in_rows, colWidths=["55%", "45%"])
        in_table.setStyle(TableStyle([
            ("BACKGROUND",   (0, 0), (-1, 0), navy),
            ("TEXTCOLOR",    (0, 0), (-1, 0), white),
            ("FONTNAME",     (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",     (0, 0), (-1, 0), 10),
            ("ALIGN",        (0, 0), (-1, 0), "LEFT"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [light, white]),
            ("FONTNAME",     (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE",     (0, 1), (-1, -1), 9),
            ("TOPPADDING",   (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 6),
            ("LEFTPADDING",  (0, 0), (-1, -1), 8),
            ("GRID",         (0, 0), (-1, -1), 0.5, colors.Color(0.8, 0.8, 0.8)),
            ("ROUNDEDCORNERS", [4]),
        ]))
        story.append(in_table)
        story.append(Spacer(1, 0.4*cm))

        # ---- Results Table ----
        if results:
            story.append(Paragraph("Calculation Results", style_h2))
            story.append(HRFlowable(width="100%", thickness=0.5, color=blue,
                                    spaceBefore=2, spaceAfter=6))

            res_rows = [["Result", "Value"]]
            for k, v in results.items():
                res_rows.append([k, str(v)])

            res_table = Table(res_rows, colWidths=["55%", "45%"])
            res_table.setStyle(TableStyle([
                ("BACKGROUND",   (0, 0), (-1, 0), navy),
                ("TEXTCOLOR",    (0, 0), (-1, 0), white),
                ("FONTNAME",     (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE",     (0, 0), (-1, 0), 10),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [light, white]),
                ("FONTNAME",     (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE",     (0, 1), (-1, -1), 9),
                ("TOPPADDING",   (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING",(0, 0), (-1, -1), 6),
                ("LEFTPADDING",  (0, 0), (-1, -1), 8),
                ("GRID",         (0, 0), (-1, -1), 0.5, colors.Color(0.8, 0.8, 0.8)),
            ]))

            # Colour the compliance result cell
            for i, (k, v) in enumerate(results.items(), start=1):
                if "Compliance" in k or "Result" in k:
                    cell_color = green if "PASS" in str(v) else red
                    res_table.setStyle(TableStyle([
                        ("BACKGROUND", (1, i), (1, i), cell_color),
                        ("TEXTCOLOR",  (1, i), (1, i), white),
                        ("FONTNAME",   (1, i), (1, i), "Helvetica-Bold"),
                    ]))

            story.append(res_table)
            story.append(Spacer(1, 0.6*cm))

        # ---- Detailed Workings (short circuit only) ----
        if hasattr(calc, "get_workings"):
            workings_text = calc.get_workings()
            if workings_text:
                story.append(Paragraph("Detailed Calculation Workings", style_h2))
                story.append(HRFlowable(width="100%", thickness=0.5, color=blue,
                                        spaceBefore=2, spaceAfter=6))
                style_mono = ParagraphStyle("mono",
                    fontSize=7.5, fontName="Courier",
                    textColor=colors.black, leading=11,
                    spaceAfter=2, leftIndent=0)
                for line in workings_text.splitlines():
                    # Preserve leading spaces by replacing with non-breaking spaces
                    display = line.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
                    leading_spaces = len(display) - len(display.lstrip(" "))
                    display = "\u00a0" * leading_spaces + display.lstrip(" ")
                    story.append(Paragraph(display or "\u00a0", style_mono))
                story.append(Spacer(1, 0.4*cm))

        # ---- Footer ----
        story.append(HRFlowable(width="100%", thickness=0.5, color=grey,
                                spaceBefore=4, spaceAfter=4))
        story.append(Paragraph(
            f"This report was generated by {Exporter.APP_NAME}  •  "
            f"{Exporter.COMPANY}  •  {now}  •  {Exporter.SANS_REF}",
            style_footer
        ))
        story.append(Paragraph(
            "This document is for engineering reference only. "
            "All designs must be verified by a registered professional.",
            ParagraphStyle("disclaimer", fontSize=7, fontName="Helvetica-Oblique",
                           textColor=grey, alignment=TA_CENTER)
        ))

        doc.build(story)

    # ------------------------------------------------------------------
    # Excel Export
    # ------------------------------------------------------------------

    @staticmethod
    def export_excel(calc: "BaseCalculator", filepath: str) -> None:
        from openpyxl import Workbook
        from openpyxl.styles import (
            PatternFill, Font, Alignment, Border, Side, GradientFill
        )
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Voltage Drop Report"

        # Column widths
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 30

        # Colours
        NAVY    = "1F3F60"
        BLUE    = "58A6FF"
        GREY    = "F4F6F8"
        WHITE   = "FFFFFF"
        GREEN   = "30B950"
        RED     = "F85149"
        LGREY   = "EEF1F4"

        def hdr_fill(hex_color):
            return PatternFill("solid", fgColor=hex_color)

        def border_thin():
            s = Side(style="thin", color="CCCCCC")
            return Border(left=s, right=s, top=s, bottom=s)

        now = datetime.now().strftime("%d %B %Y  %H:%M")

        row = 1

        # ---- App header ----
        ws.merge_cells(f"A{row}:B{row}")
        c = ws.cell(row, 1, Exporter.APP_NAME)
        c.font = Font(name="Calibri", bold=True, size=16, color=WHITE)
        c.fill = hdr_fill(NAVY)
        c.alignment = Alignment(horizontal="left", vertical="center",
                                indent=1)
        ws.row_dimensions[row].height = 30
        row += 1

        ws.merge_cells(f"A{row}:B{row}")
        c = ws.cell(row, 1, f"{Exporter.COMPANY}  |  Generated: {now}")
        c.font = Font(name="Calibri", size=9, color="888888")
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 18
        row += 2

        # ---- Calculator title ----
        ws.merge_cells(f"A{row}:B{row}")
        c = ws.cell(row, 1, calc.calculator_name)
        c.font = Font(name="Calibri", bold=True, size=13, color=NAVY)
        c.alignment = Alignment(vertical="center", indent=1)
        row += 1
        ws.merge_cells(f"A{row}:B{row}")
        c = ws.cell(row, 1, f"Reference: {calc.sans_reference}")
        c.font = Font(name="Calibri", size=9, italic=True, color="666666")
        c.alignment = Alignment(indent=1)
        row += 2

        # ---- Inputs section ----
        for col_i, header in enumerate(["Input Parameter", "Value"], 1):
            c = ws.cell(row, col_i, header)
            c.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
            c.fill = hdr_fill(NAVY)
            c.alignment = Alignment(horizontal="left", vertical="center",
                                    indent=1)
            c.border = border_thin()
            ws.row_dimensions[row].height = 22
        row += 1

        inputs = calc.get_inputs()
        for i, (k, v) in enumerate(inputs.items()):
            bg = GREY if i % 2 == 0 else WHITE
            for col_i, val in enumerate([k, str(v)], 1):
                c = ws.cell(row, col_i, val)
                c.font = Font(name="Calibri", size=10)
                c.fill = hdr_fill(bg)
                c.alignment = Alignment(horizontal="left", vertical="center",
                                        indent=1)
                c.border = border_thin()
            ws.row_dimensions[row].height = 18
            row += 1

        row += 1

        # ---- Results section ----
        results = calc.get_results()
        if results:
            for col_i, header in enumerate(["Result", "Value"], 1):
                c = ws.cell(row, col_i, header)
                c.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
                c.fill = hdr_fill(NAVY)
                c.alignment = Alignment(horizontal="left", vertical="center",
                                        indent=1)
                c.border = border_thin()
                ws.row_dimensions[row].height = 22
            row += 1

            for i, (k, v) in enumerate(results.items()):
                bg = GREY if i % 2 == 0 else WHITE
                # Highlight compliance row
                val_str = str(v)
                if "Compliance" in k or "Result" in k:
                    bg = GREEN if "PASS" in val_str else RED
                    txt_color = WHITE
                else:
                    txt_color = "000000"

                for col_i, val in enumerate([k, val_str], 1):
                    c = ws.cell(row, col_i, val)
                    c.font = Font(name="Calibri", size=10,
                                  bold=("Compliance" in k or "Result" in k),
                                  color=txt_color)
                    c.fill = hdr_fill(bg)
                    c.alignment = Alignment(horizontal="left", vertical="center",
                                            indent=1)
                    c.border = border_thin()
                ws.row_dimensions[row].height = 18
                row += 1

        row += 2

        # ---- Footer ----
        ws.merge_cells(f"A{row}:B{row}")
        c = ws.cell(row, 1,
            f"Generated by {Exporter.APP_NAME}  •  {Exporter.COMPANY}  "
            f"•  {now}  •  {Exporter.SANS_REF}")
        c.font = Font(name="Calibri", size=8, italic=True, color="AAAAAA")
        c.alignment = Alignment(horizontal="center")

        # ---- Detailed Workings sheet (short circuit only) ----
        if hasattr(calc, "get_workings"):
            workings_text = calc.get_workings()
            if workings_text:
                ws_w = wb.create_sheet("Workings")
                ws_w.column_dimensions["A"].width = 100

                ws_w.merge_cells("A1:A2")
                hdr = ws_w.cell(1, 1, "Detailed Calculation Workings")
                hdr.font = Font(name="Calibri", bold=True, size=13, color="1F3F60")
                hdr.alignment = Alignment(vertical="center", indent=1)
                ws_w.row_dimensions[1].height = 28

                for r_idx, line in enumerate(workings_text.splitlines(), start=3):
                    cell = ws_w.cell(r_idx, 1, line)
                    cell.font = Font(name="Courier New", size=9)
                    cell.alignment = Alignment(horizontal="left", vertical="center",
                                               indent=0, wrap_text=False)
                    ws_w.row_dimensions[r_idx].height = 14

        wb.save(filepath)

    # ------------------------------------------------------------------
    # Multi-Cable Comparison Table — PDF Export
    # ------------------------------------------------------------------

    @staticmethod
    def export_table_pdf(headers: list, rows: list[list], filepath: str) -> None:
        from reportlab.lib.pagesizes import A3, landscape
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.lib import colors
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
        )
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT

        doc = SimpleDocTemplate(
            filepath,
            pagesize=landscape(A3),
            leftMargin=1.5*cm, rightMargin=1.5*cm,
            topMargin=1.5*cm, bottomMargin=1.5*cm,
        )

        navy  = colors.Color(31/255, 63/255, 96/255)
        blue  = colors.Color(88/255, 166/255, 255/255)
        green = colors.Color(48/255, 185/255, 80/255)
        red   = colors.Color(248/255, 81/255, 73/255)
        grey  = colors.Color(0.55, 0.55, 0.55)
        light = colors.Color(0.94, 0.96, 0.98)
        white = colors.white

        style_h1 = ParagraphStyle("h1", fontSize=16, fontName="Helvetica-Bold",
                                   textColor=navy, spaceAfter=4)
        style_sub = ParagraphStyle("sub", fontSize=9, fontName="Helvetica",
                                    textColor=grey, spaceAfter=2)
        style_footer = ParagraphStyle("footer", fontSize=8, fontName="Helvetica",
                                       textColor=grey, alignment=TA_CENTER)

        now = datetime.now().strftime("%d %B %Y  %H:%M")
        story = []

        # Header
        story.append(Paragraph(f"{Exporter.APP_NAME}", style_h1))
        story.append(Paragraph(
            f"Multi-Cable Voltage Drop Comparison  |  {Exporter.COMPANY}  |  Generated: {now}",
            style_sub))
        story.append(HRFlowable(width="100%", thickness=2, color=navy,
                                 spaceBefore=4, spaceAfter=8))

        # Build table data
        tbl_data = [headers] + rows

        col_count = len(headers)
        # Distribute widths: small fixed cols + stretch for label
        page_w = landscape(A3)[0] - 3*cm
        w_num    = 1.2*cm
        w_result = 2.5*cm
        w_label  = page_w - w_num - w_result - 6 * 2.8*cm
        col_widths = [w_num, w_label] + [2.8*cm]*6 + [2.8*cm, w_result]

        tbl = Table(tbl_data, colWidths=col_widths, repeatRows=1)
        style_cmds = [
            # Header row
            ("BACKGROUND",    (0, 0), (-1, 0), navy),
            ("TEXTCOLOR",     (0, 0), (-1, 0), white),
            ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",      (0, 0), (-1, 0), 9),
            ("ALIGN",         (0, 0), (-1, 0), "CENTER"),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
            # Data rows
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [light, white]),
            ("FONTNAME",      (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE",      (0, 1), (-1, -1), 8),
            ("ALIGN",         (0, 1), (-1, -1), "CENTER"),
            ("ALIGN",         (1, 1), (1, -1), "LEFT"),  # label col left-aligned
            ("TOPPADDING",    (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("GRID",          (0, 0), (-1, -1), 0.4, colors.Color(0.8, 0.8, 0.8)),
        ]

        # Colour PASS/FAIL cells (last column)
        result_col = len(headers) - 1
        for row_i, row in enumerate(rows, start=1):
            val = row[result_col] if row else ""
            cell_color = green if "PASS" in val else red
            style_cmds.append(("BACKGROUND", (result_col, row_i), (result_col, row_i), cell_color))
            style_cmds.append(("TEXTCOLOR",  (result_col, row_i), (result_col, row_i), white))
            style_cmds.append(("FONTNAME",   (result_col, row_i), (result_col, row_i), "Helvetica-Bold"))

        tbl.setStyle(TableStyle(style_cmds))
        story.append(tbl)
        story.append(Spacer(1, 0.5*cm))

        # Footer
        story.append(HRFlowable(width="100%", thickness=0.5, color=grey,
                                 spaceBefore=4, spaceAfter=4))
        story.append(Paragraph(
            f"Generated by {Exporter.APP_NAME}  •  {Exporter.COMPANY}  "
            f"•  {now}  •  {Exporter.SANS_REF}  •  "
            f"Pass criterion: Vd ≤ 5%  (SANS 10142-1:2020 §6.2.7)",
            style_footer
        ))

        doc.build(story)

    # ------------------------------------------------------------------
    # Multi-Cable Comparison Table — Excel Export
    # ------------------------------------------------------------------

    @staticmethod
    def export_table_excel(headers: list, rows: list[list], filepath: str) -> None:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Cable Comparison"

        NAVY  = "1F3F60"; WHITE = "FFFFFF"; GREY = "F4F6F8"
        GREEN = "30B950"; RED   = "F85149"

        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        def fill(hex_c): return PatternFill("solid", fgColor=hex_c)

        now = datetime.now().strftime("%d %B %Y  %H:%M")
        n_cols = len(headers)

        row = 1
        # App header
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
        c = ws.cell(row, 1, Exporter.APP_NAME + " — Multi-Cable Voltage Drop Comparison")
        c.font = Font(name="Calibri", bold=True, size=14, color=WHITE)
        c.fill = fill(NAVY)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 28
        row += 1

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
        c = ws.cell(row, 1, f"{Exporter.COMPANY}  |  Generated: {now}  |  {Exporter.SANS_REF}")
        c.font = Font(name="Calibri", size=9, color="888888")
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 18
        row += 2

        # Column headers
        col_widths_map = {
            1: 6, 2: 32, 3: 18, 4: 16, 5: 24, 6: 12, 7: 14, 8: 12, 9: 10, 10: 12
        }
        for col_i, header in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(col_i)].width = col_widths_map.get(col_i, 14)
            c = ws.cell(row, col_i, header)
            c.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
            c.fill = fill(NAVY)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border
        ws.row_dimensions[row].height = 22
        row += 1

        # Data rows
        result_col = len(headers)  # last column
        for i, data_row in enumerate(rows):
            bg = GREY if i % 2 == 0 else WHITE
            for col_i, val in enumerate(data_row, 1):
                is_result = (col_i == result_col)
                if is_result:
                    cell_bg = GREEN if "PASS" in str(val) else RED
                    txt_col = WHITE
                    bold = True
                else:
                    cell_bg = bg
                    txt_col = "000000"
                    bold = False

                c = ws.cell(row, col_i, val)
                c.font = Font(name="Calibri", size=10, bold=bold, color=txt_col)
                c.fill = fill(cell_bg)
                c.alignment = Alignment(
                    horizontal="left" if col_i == 2 else "center",
                    vertical="center", indent=1 if col_i == 2 else 0
                )
                c.border = border
            ws.row_dimensions[row].height = 18
            row += 1

        row += 2
        # Footer
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
        c = ws.cell(row, 1,
            f"Generated by {Exporter.APP_NAME}  •  {Exporter.COMPANY}  "
            f"•  {now}  •  Pass criterion: Vd ≤ 5%  (SANS 10142-1:2020 §6.2.7)")
        c.font = Font(name="Calibri", size=8, italic=True, color="AAAAAA")
        c.alignment = Alignment(horizontal="center")

        wb.save(filepath)
