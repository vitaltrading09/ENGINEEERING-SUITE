"""
gland_size_widget.py
--------------------
Cable Gland Size Calculator.

Features:
  - Per-cable individual End-A / End-B gland toggle
  - Bulk selection controls (All A, All B, All, None)
  - Separate 3-core and 4-core OD entries
  - CCG BW (SWA) / CCG A2 (H07) automatic type selection
  - Optional lug BOQ section
  - PDF + Excel export
"""

from __future__ import annotations

from datetime import datetime

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QComboBox, QSpinBox,
    QGroupBox, QPushButton, QFrame, QSizePolicy, QFileDialog, QMessageBox,
    QTableWidget, QTableWidgetItem, QHeaderView, QCheckBox, QLineEdit,
    QScrollArea,
)
from PyQt6.QtCore import Qt, QSize
from PyQt6.QtGui import QFont, QPainter, QColor, QPen, QBrush, QPainterPath
import qtawesome as qta

from calculators.base_calculator import BaseCalculator
from calculators.gland_size.gland_size_logic import (
    calc_gland, find_gland,
    CABLE_OD, CABLE_GLAND_TYPE, CABLE_CORES,
    LUG_SIZES_MM2, LUG_BOLT_SIZES,
    GlandCalcResult, GlandMatch, BOQItem,
)


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _sep() -> QFrame:
    f = QFrame()
    f.setFrameShape(QFrame.Shape.HLine)
    f.setFrameShadow(QFrame.Shadow.Sunken)
    return f


def _note(text: str) -> QLabel:
    lbl = QLabel(text)
    lbl.setObjectName("header_subtitle")
    lbl.setWordWrap(True)
    return lbl


def _input_row(label: str, widget: QWidget, unit: str = "") -> QHBoxLayout:
    row = QHBoxLayout()
    lbl = QLabel(label)
    lbl.setFixedWidth(200)
    row.addWidget(lbl)
    row.addWidget(widget, 1)
    if unit:
        u = QLabel(unit)
        u.setObjectName("result_label")
        u.setFixedWidth(60)
        row.addWidget(u)
    else:
        row.addSpacing(60)
    return row


def _action_btn(text: str, icon_name: str, obj_name: str, icon_color: str, height: int = 42):
    btn = QPushButton(text)
    btn.setObjectName(obj_name)
    btn.setIcon(qta.icon(icon_name, color=icon_color))
    btn.setIconSize(QSize(14, 14))
    btn.setFixedHeight(height)
    btn.setCursor(Qt.CursorShape.PointingHandCursor)
    return btn


_TOGGLE_SS = """
QPushButton {{
    border-radius: 16px;
    font-weight: bold;
    font-size: 10px;
    border: 2px solid {rim};
    background: {bg};
    color: {fg};
}}
QPushButton:hover {{ border-color: #58a6ff; }}
"""

_TOGGLE_ON  = _TOGGLE_SS.format(rim="#30b950", bg="#30b950", fg="white")
_TOGGLE_OFF = _TOGGLE_SS.format(rim="#484e57", bg="#3c4148", fg="#8b949e")


# ─────────────────────────────────────────────────────────────────────────────
# End toggle button
# ─────────────────────────────────────────────────────────────────────────────

class EndToggle(QPushButton):
    """32×32 circular checkable button — green=gland required, grey=none."""

    def __init__(self, end_label: str, parent=None):
        super().__init__(parent)
        self._label = end_label
        self.setCheckable(True)
        self.setChecked(False)
        self.setFixedSize(32, 32)
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        self._refresh()
        self.toggled.connect(lambda _: self._refresh())

    def _refresh(self):
        if self.isChecked():
            self.setText("G")
            self.setStyleSheet(_TOGGLE_ON)
            self.setToolTip(f"End {self._label}: Gland required — click to remove")
        else:
            self.setText(self._label)
            self.setStyleSheet(_TOGGLE_OFF)
            self.setToolTip(f"End {self._label}: No gland — click to add gland")


# ─────────────────────────────────────────────────────────────────────────────
# Cable body painter
# ─────────────────────────────────────────────────────────────────────────────

class _CableBody(QWidget):
    """Draws a simple cable cross-section strip."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setMinimumWidth(60)
        self.setFixedHeight(32)
        self.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)

    def paintEvent(self, event):
        p = QPainter(self)
        p.setRenderHint(QPainter.RenderHint.Antialiasing)
        w, h = self.width(), self.height()

        # Sheath
        path = QPainterPath()
        path.addRoundedRect(0, 4, w, h - 8, 4, 4)
        p.setPen(Qt.PenStyle.NoPen)
        p.setBrush(QBrush(QColor(45, 52, 62)))
        p.drawPath(path)

        # Insulation stripe
        inner = QPainterPath()
        inner.addRoundedRect(3, 7, w - 6, h - 14, 3, 3)
        p.setBrush(QBrush(QColor(58, 100, 148)))
        p.drawPath(inner)

        # Highlight
        p.setBrush(QBrush(QColor(255, 255, 255, 20)))
        hi = QPainterPath()
        hi.addRoundedRect(6, 8, w - 12, 4, 2, 2)
        p.drawPath(hi)

        # End connectors (short vertical tabs)
        p.setBrush(QBrush(QColor(88, 166, 255, 180)))
        p.drawRect(0, 6, 4, h - 12)
        p.drawRect(w - 4, 6, 4, h - 12)


# ─────────────────────────────────────────────────────────────────────────────
# Single cable row
# ─────────────────────────────────────────────────────────────────────────────

class CableRow(QWidget):
    """One cable run with independent End-A and End-B gland toggles."""

    def __init__(self, index: int, parent=None):
        super().__init__(parent)
        self.index = index

        lay = QHBoxLayout(self)
        lay.setContentsMargins(4, 2, 4, 2)
        lay.setSpacing(6)

        # Cable number label
        num = QLabel(f"{index + 1}")
        num.setFixedWidth(20)
        num.setAlignment(Qt.AlignmentFlag.AlignCenter)
        num.setObjectName("header_subtitle")
        lay.addWidget(num)

        # End A toggle
        a_lbl = QLabel("A")
        a_lbl.setFixedWidth(10)
        a_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        a_lbl.setObjectName("header_subtitle")
        self.btn_a = EndToggle("A")
        lay.addWidget(a_lbl)
        lay.addWidget(self.btn_a)

        # Cable body
        lay.addSpacing(4)
        self._body = _CableBody()
        lay.addWidget(self._body, 1)
        lay.addSpacing(4)

        # End B toggle
        self.btn_b = EndToggle("B")
        b_lbl = QLabel("B")
        b_lbl.setFixedWidth(10)
        b_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        b_lbl.setObjectName("header_subtitle")
        lay.addWidget(self.btn_b)
        lay.addWidget(b_lbl)

        self.setFixedHeight(42)

    def get_selection(self) -> tuple[bool, bool]:
        return (self.btn_a.isChecked(), self.btn_b.isChecked())

    def set_selection(self, a: bool, b: bool):
        self.btn_a.setChecked(a)
        self.btn_b.setChecked(b)


# ─────────────────────────────────────────────────────────────────────────────
# Cable run panel (scrollable list of CableRows)
# ─────────────────────────────────────────────────────────────────────────────

class CableRunPanel(QWidget):
    """
    Displays one CableRow per cable. Shows End-A and End-B toggles per row.
    """

    def __init__(self, parent=None):
        super().__init__(parent)
        self._rows: list[CableRow] = []

        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(6)

        # ── Column header ─────────────────────────────────────────────────
        hdr = QHBoxLayout()
        hdr.setContentsMargins(4, 0, 4, 0)
        hdr.addSpacing(20)                             # cable number col
        hdr.addSpacing(10 + 32 + 6)                   # A label + toggle
        hdr.addSpacing(4)
        center_lbl = QLabel("Cable Run")
        center_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        center_lbl.setObjectName("header_subtitle")
        hdr.addWidget(center_lbl, 1)
        hdr.addSpacing(4)
        hdr.addSpacing(32 + 6 + 10)                   # toggle + B label
        outer.addLayout(hdr)

        outer.addWidget(_sep())

        # ── Bulk selection bar ────────────────────────────────────────────
        bulk = QHBoxLayout()
        bulk.setContentsMargins(4, 0, 4, 0)
        bulk.setSpacing(6)
        bulk.addWidget(QLabel("Select:"))
        for label, slot in (
            ("All A",    self._select_all_a),
            ("All B",    self._select_all_b),
            ("All",      self._select_all),
            ("None",     self._select_none),
        ):
            btn = QPushButton(label)
            btn.setFixedHeight(26)
            btn.setObjectName("secondary_btn")
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            btn.clicked.connect(slot)
            bulk.addWidget(btn)
        bulk.addStretch()
        outer.addLayout(bulk)

        # ── Scrollable rows ───────────────────────────────────────────────
        self._scroll = QScrollArea()
        self._scroll.setWidgetResizable(True)
        self._scroll.setMinimumHeight(260)
        self._scroll.setMaximumHeight(520)
        self._scroll.setHorizontalScrollBarPolicy(
            Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self._scroll.setStyleSheet(
            "QScrollArea { border: none; background: transparent; }")

        self._container = QWidget()
        self._row_layout = QVBoxLayout(self._container)
        self._row_layout.setContentsMargins(0, 2, 0, 2)
        self._row_layout.setSpacing(3)
        self._row_layout.addStretch()
        self._scroll.setWidget(self._container)
        outer.addWidget(self._scroll)

    # ── Public API ────────────────────────────────────────────────────────

    def set_qty(self, qty: int):
        """Add or remove rows to match the requested cable count."""
        current = len(self._rows)

        if qty > current:
            for i in range(current, qty):
                row = CableRow(i)
                self._rows.append(row)
                self._row_layout.insertWidget(self._row_layout.count() - 1, row)
        elif qty < current:
            for row in self._rows[qty:]:
                self._row_layout.removeWidget(row)
                row.deleteLater()
            self._rows = self._rows[:qty]

    def get_selections(self) -> list[tuple[bool, bool]]:
        return [r.get_selection() for r in self._rows]

    # ── Bulk controls ─────────────────────────────────────────────────────

    def _select_all_a(self):
        for r in self._rows:
            r.btn_a.setChecked(True)

    def _select_all_b(self):
        for r in self._rows:
            r.btn_b.setChecked(True)

    def _select_all(self):
        for r in self._rows:
            r.btn_a.setChecked(True)
            r.btn_b.setChecked(True)

    def _select_none(self):
        for r in self._rows:
            r.btn_a.setChecked(False)
            r.btn_b.setChecked(False)


# ─────────────────────────────────────────────────────────────────────────────
# Gland result card
# ─────────────────────────────────────────────────────────────────────────────

class GlandCard(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        lay = QVBoxLayout(self)
        lay.setContentsMargins(12, 10, 12, 10)
        lay.setSpacing(4)

        hdr = QHBoxLayout()
        ico = QLabel()
        ico.setPixmap(qta.icon("fa5s.plug", color="#58a6ff").pixmap(14, 14))
        hdr.addWidget(ico)
        hdr.addSpacing(4)
        title = QLabel("Recommended Gland")
        title.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
        title.setObjectName("section_title")
        hdr.addWidget(title)
        hdr.addStretch()
        self.badge = QLabel("—")
        self.badge.setObjectName("neutral_badge")
        self.badge.setFixedHeight(22)
        hdr.addWidget(self.badge)
        lay.addLayout(hdr)

        self.prod_lbl   = QLabel("Product Code:  —")
        self.size_lbl   = QLabel("Size Reference: —")
        self.thread_lbl = QLabel("Thread:         —")
        self.od_lbl     = QLabel("OD Range:       —")
        self.count_lbl  = QLabel("Total glands:   —")
        for lbl in (self.prod_lbl, self.size_lbl, self.thread_lbl,
                    self.od_lbl, self.count_lbl):
            lbl.setObjectName("header_subtitle")
            lay.addWidget(lbl)

        self.setObjectName("result_card")
        self.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)

    def set_result(self, result: GlandCalcResult):
        m = result.gland_match
        total = result.total_glands_a + result.total_glands_b
        if total == 0:
            self.badge.setText("NONE SEL.")
            self.badge.setObjectName("neutral_badge")
            self.prod_lbl.setText("No gland ends selected.")
            self.size_lbl.setText("")
            self.thread_lbl.setText("")
            self.od_lbl.setText("")
            self.count_lbl.setText("")
        elif m and m.matched:
            self.badge.setText("MATCH")
            self.badge.setObjectName("pass_badge")
            self.prod_lbl.setText(f"Product Code:   {m.product_code}")
            self.size_lbl.setText(f"Size Reference: {m.size_ref}")
            self.thread_lbl.setText(f"Thread:         {m.thread}")
            self.od_lbl.setText(
                f"OD Range:       {m.od_min:.1f} – {m.od_max:.1f} mm"
                f"  (cable OD = {m.cable_od:.1f} mm)")
            self.count_lbl.setText(
                f"Total glands:   {total}  "
                f"({result.total_glands_a}× End-A  +  {result.total_glands_b}× End-B)")
        else:
            self.badge.setText("NO MATCH")
            self.badge.setObjectName("fail_badge")
            od_txt = f"{result.cable_od:.1f} mm" if result.cable_od > 0 else "unknown OD"
            self.prod_lbl.setText(f"No gland found for OD {od_txt}")
            self.size_lbl.setText("Cable OD is outside CCG database range.")
            self.thread_lbl.setText("Contact CCG for a special size.")
            self.od_lbl.setText("")
            self.count_lbl.setText(f"Total glands needed: {total}")
        self.badge.style().unpolish(self.badge)
        self.badge.style().polish(self.badge)


# ─────────────────────────────────────────────────────────────────────────────
# Main widget
# ─────────────────────────────────────────────────────────────────────────────

class GlandSizeWidget(BaseCalculator):
    calculator_name = "Cable Gland Size & BOQ"
    sans_reference  = "CCG GI010214 / GI020816  •  SANS 1213  •  BS 6121"

    def __init__(self, parent=None):
        super().__init__(parent)
        self._result: GlandCalcResult | None = None
        self._build_ui()

    # ─────────────────────────────────────────────────────── UI Build ─────

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(32, 28, 32, 32)
        root.setSpacing(20)

        # ── Title ─────────────────────────────────────────────────────────
        tr = QHBoxLayout()
        ico = QLabel()
        ico.setPixmap(qta.icon("fa5s.plug", color="#58a6ff").pixmap(28, 28))
        tr.addWidget(ico)
        tr.addSpacing(10)
        tc = QVBoxLayout()
        tc.setSpacing(2)
        h1 = QLabel("Cable Gland Size & BOQ")
        h1.setFont(QFont("Segoe UI", 18, QFont.Weight.Bold))
        h1.setObjectName("section_title")
        tc.addWidget(h1)
        tc.addWidget(_note(
            "CCG BW Captive Cone Gland (SWA)  |  CCG A2 Compression Gland (H07RN-F)  |  "
            "SANS 1213 / BS 6121  |  Toggle each cable end individually"
        ))
        tr.addLayout(tc)
        tr.addStretch()
        root.addLayout(tr)
        root.addWidget(_sep())

        # ── Two-column body ───────────────────────────────────────────────
        cols = QHBoxLayout()
        cols.setSpacing(24)
        left  = QVBoxLayout()
        left.setSpacing(14)
        right = QVBoxLayout()
        right.setSpacing(14)

        # ══════════════════════════════════════ LEFT COLUMN ═══════════════

        # ── Cable Configuration ───────────────────────────────────────────
        cg = QGroupBox("Cable Configuration")
        cl = QVBoxLayout(cg)
        cl.setSpacing(10)
        cl.addWidget(_note(
            "Select cable type (3c = 3-core, 4c = 4-core — they have different ODs). "
            "BW glands are used for SWA armoured; A2 compression for H07RN-F."
        ))

        self.cable_type_combo = QComboBox()
        for ct in CABLE_OD:
            self.cable_type_combo.addItem(ct, userData=ct)
        self.cable_type_combo.currentIndexChanged.connect(self._on_cable_type_changed)
        cl.addLayout(_input_row("Cable Type", self.cable_type_combo))

        self.cable_size_combo = QComboBox()
        self.cable_size_combo.currentIndexChanged.connect(self._on_cable_size_changed)
        cl.addLayout(_input_row("Cable Size", self.cable_size_combo, "mm²"))

        self.cable_od_lbl = QLabel("OD = —")
        self.cable_od_lbl.setObjectName("header_subtitle")
        cl.addWidget(self.cable_od_lbl)

        self.cable_qty_spin = QSpinBox()
        self.cable_qty_spin.setRange(1, 30)
        self.cable_qty_spin.setValue(1)
        self.cable_qty_spin.valueChanged.connect(self._on_qty_changed)
        cl.addLayout(_input_row("Quantity", self.cable_qty_spin, "cables"))

        self.run_label_edit = QLineEdit()
        self.run_label_edit.setPlaceholderText(
            "e.g.  C01 — MDB to DB-1  (optional)")
        cl.addLayout(_input_row("Run Label", self.run_label_edit))

        left.addWidget(cg)

        # ── Per-cable diagram ─────────────────────────────────────────────
        dg = QGroupBox("Cable Run  —  Click G to Toggle Gland at Each End")
        dl = QVBoxLayout(dg)
        dl.setSpacing(6)
        dl.addWidget(_note(
            "Each row is one cable.  "
            "Green [G] = gland required at that end.  "
            "Grey [A/B] = no gland.  "
            "Use the bulk buttons to select all ends at once."
        ))
        self.run_panel = CableRunPanel()
        dl.addWidget(self.run_panel)
        left.addWidget(dg)

        # ── Lug Selection ─────────────────────────────────────────────────
        lg = QGroupBox("Lug Selection  (optional)")
        ll = QVBoxLayout(lg)
        ll.setSpacing(10)

        self.lug_enable_chk = QCheckBox("Include cable lugs in BOQ")
        self.lug_enable_chk.toggled.connect(self._toggle_lug_panel)
        ll.addWidget(self.lug_enable_chk)

        self._lug_panel = QWidget()
        lpl = QVBoxLayout(self._lug_panel)
        lpl.setContentsMargins(0, 4, 0, 0)
        lpl.setSpacing(8)
        lpl.addWidget(_note(
            "Qty = No. of gland ends selected × No. of cores. "
            "Conductor size auto-matches cable size; adjust if using reduced lugs."
        ))

        self.lug_size_combo = QComboBox()
        for s in LUG_SIZES_MM2:
            self.lug_size_combo.addItem(f"{s} mm²", userData=float(s))
        lpl.addLayout(_input_row("Conductor Size", self.lug_size_combo, "mm²"))

        self.lug_bolt_combo = QComboBox()
        for b in LUG_BOLT_SIZES:
            self.lug_bolt_combo.addItem(b, userData=b)
        self.lug_bolt_combo.setCurrentIndex(3)   # M10 default
        lpl.addLayout(_input_row("Bolt Hole", self.lug_bolt_combo))

        self.lug_cores_spin = QSpinBox()
        self.lug_cores_spin.setRange(1, 5)
        self.lug_cores_spin.setValue(3)
        lpl.addLayout(_input_row("Cores / Lugs per end", self.lug_cores_spin))

        self._lug_panel.setVisible(False)
        ll.addWidget(self._lug_panel)
        left.addWidget(lg)

        # ── Action buttons ────────────────────────────────────────────────
        brow = QHBoxLayout()
        brow.setSpacing(10)
        self.calc_btn  = _action_btn("  Calculate",  "fa5s.calculator",  "primary_btn",   "white")
        self.pdf_btn   = _action_btn("  PDF",        "fa5s.file-pdf",    "secondary_btn", "#8b949e")
        self.excel_btn = _action_btn("  Excel",      "fa5s.file-excel",  "secondary_btn", "#8b949e")
        self.reset_btn = _action_btn("  Reset",      "fa5s.redo",        "danger_btn",    "#f85149")
        self.calc_btn.clicked.connect(self._calculate)
        self.pdf_btn.clicked.connect(self._export_pdf)
        self.excel_btn.clicked.connect(self._export_excel)
        self.reset_btn.clicked.connect(self.reset)
        for b in (self.calc_btn, self.pdf_btn, self.excel_btn, self.reset_btn):
            brow.addWidget(b)
        left.addLayout(brow)
        left.addStretch()

        # ══════════════════════════════════════ RIGHT COLUMN ══════════════

        # ── Gland result card ─────────────────────────────────────────────
        self.gland_card = GlandCard()
        right.addWidget(self.gland_card)

        # ── BOQ Table ─────────────────────────────────────────────────────
        bq = QGroupBox("Bill of Quantities (BOQ)")
        bql = QVBoxLayout(bq)
        bql.setSpacing(6)
        bql.addWidget(_note(
            "Quantities are calculated from individual end selections × cores."
        ))
        self.boq_table = QTableWidget()
        self.boq_table.setColumnCount(5)
        self.boq_table.setHorizontalHeaderLabels(
            ["No.", "Description", "Product Code", "Qty", "Unit"])
        hh = self.boq_table.horizontalHeader()
        hh.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        hh.setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
        self.boq_table.setColumnWidth(0, 40)
        self.boq_table.setColumnWidth(2, 120)
        self.boq_table.setColumnWidth(3, 50)
        self.boq_table.setColumnWidth(4, 50)
        self.boq_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.boq_table.setSelectionBehavior(
            QTableWidget.SelectionBehavior.SelectRows)
        self.boq_table.setAlternatingRowColors(True)
        self.boq_table.setMinimumHeight(180)
        self.boq_table.verticalHeader().setVisible(False)
        bql.addWidget(self.boq_table)
        right.addWidget(bq)
        right.addStretch()

        cols.addLayout(left,  48)
        cols.addLayout(right, 52)
        root.addLayout(cols)

        # ── Init ──────────────────────────────────────────────────────────
        self._on_cable_type_changed()
        self._on_qty_changed(1)

    # ─────────────────────────────────────────────── Event handlers ────────

    def _on_cable_type_changed(self):
        cable_type = self.cable_type_combo.currentData()
        od_table   = CABLE_OD.get(cable_type, {})
        sizes      = sorted(od_table.keys())

        self.cable_size_combo.blockSignals(True)
        self.cable_size_combo.clear()
        for s in sizes:
            label = f"{s:.0f}" if s == int(s) else str(s)
            self.cable_size_combo.addItem(label, userData=s)
        self.cable_size_combo.blockSignals(False)

        # Auto-fill lug cores from cable type
        self.lug_cores_spin.setValue(CABLE_CORES.get(cable_type, 3))

        self._on_cable_size_changed()

    def _on_cable_size_changed(self):
        cable_type = self.cable_type_combo.currentData()
        size       = self.cable_size_combo.currentData() or 0.0
        od         = CABLE_OD.get(cable_type, {}).get(size, 0.0)
        gland_type = CABLE_GLAND_TYPE.get(cable_type, "BW")

        if od > 0:
            match = find_gland(od, gland_type)
            if match.matched:
                self.cable_od_lbl.setText(
                    f"OD ≈ {od:.1f} mm   |   Gland type: CCG {gland_type}   |   "
                    f"→ {match.size_ref}  ({match.thread})")
            else:
                self.cable_od_lbl.setText(
                    f"OD ≈ {od:.1f} mm   |   CCG {gland_type}   |   "
                    f"WARNING: no gland in range")
        else:
            self.cable_od_lbl.setText("OD not in database")

        # Sync lug size
        if size > 0:
            nearest = min(LUG_SIZES_MM2, key=lambda x: abs(x - size))
            for i in range(self.lug_size_combo.count()):
                if self.lug_size_combo.itemData(i) == float(nearest):
                    self.lug_size_combo.setCurrentIndex(i)
                    break

    def _on_qty_changed(self, qty: int):
        self.run_panel.set_qty(qty)

    def _toggle_lug_panel(self, enabled: bool):
        self._lug_panel.setVisible(enabled)

    # ──────────────────────────────────────────────── Calculation ──────────

    def _calculate(self):
        cable_type = self.cable_type_combo.currentData()
        cable_size = self.cable_size_combo.currentData() or 0.0
        selections = self.run_panel.get_selections()

        result = calc_gland(
            cable_type    = cable_type,
            cable_size_mm2= cable_size,
            selections    = selections,
            include_lugs  = self.lug_enable_chk.isChecked(),
            lug_size_mm2  = self.lug_size_combo.currentData() or 0.0,
            lug_bolt      = self.lug_bolt_combo.currentData() or "",
            lug_cores     = self.lug_cores_spin.value(),
        )
        self._result = result

        # Update gland card
        self.gland_card.set_result(result)

        # Populate BOQ table
        self.boq_table.setRowCount(len(result.boq))
        for row, item in enumerate(result.boq):
            self.boq_table.setItem(row, 0, QTableWidgetItem(str(item.item_no)))
            self.boq_table.setItem(row, 1, QTableWidgetItem(item.description))
            self.boq_table.setItem(row, 2, QTableWidgetItem(item.product_code))
            qty_cell = QTableWidgetItem(str(item.qty))
            qty_cell.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.boq_table.setItem(row, 3, qty_cell)
            unit_cell = QTableWidgetItem(item.unit)
            unit_cell.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.boq_table.setItem(row, 4, unit_cell)

    # ──────────────────────────────────────────────── Exports ──────────────

    def _export_pdf(self):
        if not self._result:
            QMessageBox.information(self, "No Results", "Run a calculation first.")
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Save PDF", "GlandSizing_Report.pdf", "PDF Files (*.pdf)")
        if path:
            try:
                self._write_pdf(path)
                QMessageBox.information(self, "Saved", f"PDF saved:\n{path}")
            except Exception as e:
                QMessageBox.critical(self, "Error", str(e))

    def _export_excel(self):
        if not self._result:
            QMessageBox.information(self, "No Results", "Run a calculation first.")
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Save Excel", "GlandSizing_Report.xlsx", "Excel Files (*.xlsx)")
        if path:
            try:
                self._write_excel(path)
                QMessageBox.information(self, "Saved", f"Excel saved:\n{path}")
            except Exception as e:
                QMessageBox.critical(self, "Error", str(e))

    def _write_pdf(self, filepath: str):
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.lib import colors
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable,
        )
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT

        r   = self._result
        now = datetime.now().strftime("%d %B %Y  %H:%M")

        navy  = colors.Color(31/255, 63/255, 96/255)
        blue  = colors.Color(88/255, 166/255, 255/255)
        green = colors.Color(48/255, 185/255, 80/255)
        red   = colors.Color(248/255, 81/255, 73/255)
        grey  = colors.Color(0.55, 0.55, 0.55)
        light = colors.Color(0.94, 0.96, 0.98)
        white = colors.white

        sh1  = ParagraphStyle("h1", fontSize=18, fontName="Helvetica-Bold",
                              textColor=navy, spaceAfter=4)
        sh2  = ParagraphStyle("h2", fontSize=12, fontName="Helvetica-Bold",
                              textColor=navy, spaceBefore=12, spaceAfter=6)
        sbody= ParagraphStyle("body", fontSize=10, fontName="Helvetica",
                              textColor=colors.black, spaceAfter=4, leading=14)
        sftr = ParagraphStyle("ftr", fontSize=8, fontName="Helvetica",
                              textColor=grey, alignment=TA_CENTER)

        tbl_style = TableStyle([
            ("BACKGROUND",    (0, 0), (-1, 0), navy),
            ("TEXTCOLOR",     (0, 0), (-1, 0), white),
            ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",      (0, 0), (-1, 0), 10),
            ("ROWBACKGROUNDS",(0, 1), (-1, -1), [light, white]),
            ("FONTNAME",      (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE",      (0, 1), (-1, -1), 9),
            ("TOPPADDING",    (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("LEFTPADDING",   (0, 0), (-1, -1), 8),
            ("GRID",          (0, 0), (-1, -1), 0.5, colors.Color(0.8, 0.8, 0.8)),
        ])

        doc = SimpleDocTemplate(
            filepath, pagesize=A4,
            leftMargin=2.2*cm, rightMargin=2.2*cm,
            topMargin=2.2*cm,  bottomMargin=2.2*cm,
        )
        story = []

        # Header
        hdr_data = [[
            Paragraph("<b>Engineering Calculator Suite</b>", sh1),
            Paragraph(
                f"<font color='grey' size='9'>Solareff Engineering<br/>"
                f"Generated: {now}</font>",
                ParagraphStyle("hr", fontSize=9, alignment=TA_RIGHT,
                               textColor=grey, fontName="Helvetica")),
        ]]
        ht = Table(hdr_data, colWidths=["65%", "35%"])
        ht.setStyle(TableStyle([
            ("VALIGN",       (0,0),(-1,-1),"MIDDLE"),
            ("LEFTPADDING",  (0,0),(-1,-1),0),
            ("RIGHTPADDING", (0,0),(-1,-1),0),
        ]))
        story.append(ht)
        story.append(HRFlowable(width="100%", thickness=2, color=navy,
                                spaceAfter=6, spaceBefore=4))
        story.append(Paragraph(self.calculator_name, sh2))
        story.append(Paragraph(
            f"Reference: <b>{self.sans_reference}</b>", sbody))
        story.append(Spacer(1, 0.3*cm))

        # Cable details
        run_label = self.run_label_edit.text().strip()
        story.append(Paragraph("Cable Details", sh2))
        story.append(HRFlowable(width="100%", thickness=0.5, color=blue,
                                spaceBefore=2, spaceAfter=6))
        info_rows = [["Parameter", "Value"]]
        if run_label:
            info_rows.append(["Run Label", run_label])
        info_rows += [
            ["Cable Type",       r.cable_type],
            ["Cable Size",       f"{r.cable_size_mm2:.0f} mm²"],
            ["Cable OD",         f"{r.cable_od:.1f} mm" if r.cable_od > 0 else "—"],
            ["Total Cables",     str(r.cable_qty)],
            ["Gland Type",       f"CCG {r.gland_type}"],
            ["Glands — End A",   str(r.total_glands_a)],
            ["Glands — End B",   str(r.total_glands_b)],
            ["Total Glands",     str(r.total_glands_a + r.total_glands_b)],
        ]
        it = Table(info_rows, colWidths=["45%", "55%"])
        it.setStyle(tbl_style)
        story.append(it)
        story.append(Spacer(1, 0.3*cm))

        # Per-cable selection table
        story.append(Paragraph("Cable-by-Cable Gland Selection", sh2))
        story.append(HRFlowable(width="100%", thickness=0.5, color=blue,
                                spaceBefore=2, spaceAfter=6))
        sel_rows = [["Cable", "End A", "End B"]]
        for i, (a, b) in enumerate(r.selections):
            sel_rows.append([
                str(i + 1),
                "Gland" if a else "—",
                "Gland" if b else "—",
            ])
        sel_tbl = Table(sel_rows, colWidths=["20%", "40%", "40%"])
        sel_tbl.setStyle(tbl_style)
        for i, (a, b) in enumerate(r.selections, start=1):
            if a:
                sel_tbl.setStyle(TableStyle([
                    ("BACKGROUND", (1, i), (1, i), green),
                    ("TEXTCOLOR",  (1, i), (1, i), white),
                    ("FONTNAME",   (1, i), (1, i), "Helvetica-Bold"),
                ]))
            if b:
                sel_tbl.setStyle(TableStyle([
                    ("BACKGROUND", (2, i), (2, i), green),
                    ("TEXTCOLOR",  (2, i), (2, i), white),
                    ("FONTNAME",   (2, i), (2, i), "Helvetica-Bold"),
                ]))
        story.append(sel_tbl)
        story.append(Spacer(1, 0.3*cm))

        # Gland recommendation
        story.append(Paragraph("Recommended Gland", sh2))
        story.append(HRFlowable(width="100%", thickness=0.5, color=blue,
                                spaceBefore=2, spaceAfter=6))
        m = r.gland_match
        if m and m.matched:
            g_rows = [["Parameter", "Value"],
                      ["Product Code",   m.product_code],
                      ["Size Reference", m.size_ref],
                      ["Thread",         m.thread],
                      ["Cable OD Range", f"{m.od_min:.1f} – {m.od_max:.1f} mm"],
                      ["Cable OD",       f"{m.cable_od:.1f} mm"],
                      ["Compliance",     "PASS — Gland fits cable OD"]]
            gt = Table(g_rows, colWidths=["45%", "55%"])
            gt.setStyle(tbl_style)
            gt.setStyle(TableStyle([("BACKGROUND",(1,6),(1,6),green),
                                    ("TEXTCOLOR", (1,6),(1,6),white),
                                    ("FONTNAME",  (1,6),(1,6),"Helvetica-Bold")]))
            story.append(gt)
        elif r.total_glands_a + r.total_glands_b > 0:
            story.append(Paragraph(
                f"<font color='red'><b>NO MATCH</b> — Cable OD "
                f"{r.cable_od:.1f} mm is outside the available CCG range. "
                f"Consult CCG for a special size.</font>", sbody))
        else:
            story.append(Paragraph("No gland ends selected.", sbody))
        story.append(Spacer(1, 0.3*cm))

        # BOQ
        if r.boq:
            story.append(Paragraph("Bill of Quantities", sh2))
            story.append(HRFlowable(width="100%", thickness=0.5, color=blue,
                                    spaceBefore=2, spaceAfter=6))
            boq_rows = [["No.", "Description", "Prod. Code", "Qty", "Unit"]]
            for item in r.boq:
                boq_rows.append([str(item.item_no), item.description,
                                  item.product_code, str(item.qty), item.unit])
            bt = Table(boq_rows, colWidths=["5%", "55%", "20%", "8%", "12%"])
            bt.setStyle(TableStyle([
                ("BACKGROUND",    (0,0),(-1,0),navy),
                ("TEXTCOLOR",     (0,0),(-1,0),white),
                ("FONTNAME",      (0,0),(-1,0),"Helvetica-Bold"),
                ("FONTSIZE",      (0,0),(-1,0),9),
                ("ROWBACKGROUNDS",(0,1),(-1,-1),[light,white]),
                ("FONTNAME",      (0,1),(-1,-1),"Helvetica"),
                ("FONTSIZE",      (0,1),(-1,-1),8),
                ("TOPPADDING",    (0,0),(-1,-1),6),
                ("BOTTOMPADDING", (0,0),(-1,-1),6),
                ("LEFTPADDING",   (0,0),(-1,-1),6),
                ("ALIGN",         (0,0),(0,-1),"CENTER"),
                ("ALIGN",         (3,0),(4,-1),"CENTER"),
                ("GRID",          (0,0),(-1,-1),0.5,colors.Color(0.8,0.8,0.8)),
            ]))
            story.append(bt)
            story.append(Spacer(1, 0.6*cm))

        story.append(HRFlowable(width="100%", thickness=0.5, color=grey,
                                spaceBefore=4, spaceAfter=4))
        story.append(Paragraph(
            f"Engineering Calculator Suite  •  Solareff Engineering  •  {now}  •  "
            "CCG GI010214 / GI020816", sftr))
        story.append(Paragraph(
            "For engineering reference only. Verify OD against actual cable before ordering.",
            ParagraphStyle("disc", fontSize=7, fontName="Helvetica-Oblique",
                           textColor=grey, alignment=TA_CENTER)))
        doc.build(story)

    def _write_excel(self, filepath: str):
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

        r   = self._result
        now = datetime.now().strftime("%d %B %Y  %H:%M")
        NAVY="1F3F60"; WHITE="FFFFFF"; LIGHT="F0F4F8"
        GREEN="30B950"; RED="F85149"

        thin = Side(style="thin", color="CCCCCC")
        bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
        def fill(h): return PatternFill("solid", fgColor=h)

        wb = Workbook()
        ws = wb.active
        ws.title = "Gland Sizing"
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 55
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 10
        row = 1

        def section_hdr(text):
            nonlocal row
            ws.merge_cells(f"A{row}:E{row}")
            c = ws.cell(row, 1, text)
            c.font = Font(name="Calibri", bold=True, size=11, color=WHITE)
            c.fill = fill(NAVY)
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws.row_dimensions[row].height = 22
            row += 1

        def data_row_2col(label, value, bg="F4F6F8"):
            nonlocal row
            for ci, val in enumerate([label, value], 1):
                c = ws.cell(row, ci, val)
                c.font = Font(name="Calibri", size=10)
                c.fill = fill(bg if ci == 1 else WHITE)
                c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                c.border = bdr
            ws.merge_cells(f"B{row}:E{row}")
            ws.row_dimensions[row].height = 18
            row += 1

        # App header
        ws.merge_cells(f"A{row}:E{row}")
        c = ws.cell(row, 1, "Engineering Calculator Suite — Cable Gland Size & BOQ")
        c.font = Font(name="Calibri", bold=True, size=14, color=WHITE)
        c.fill = fill(NAVY)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 30
        row += 1
        ws.merge_cells(f"A{row}:E{row}")
        c = ws.cell(row, 1, f"Solareff Engineering  |  Generated: {now}")
        c.font = Font(name="Calibri", size=9, color="888888")
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 18
        row += 2

        # Cable details
        section_hdr("Cable Details")
        run_label = self.run_label_edit.text().strip()
        if run_label:
            data_row_2col("Run Label", run_label)
        data_row_2col("Cable Type",     r.cable_type)
        data_row_2col("Cable Size",     f"{r.cable_size_mm2:.0f} mm²")
        data_row_2col("Cable OD",       f"{r.cable_od:.1f} mm" if r.cable_od > 0 else "—")
        data_row_2col("Total Cables",   str(r.cable_qty))
        data_row_2col("Gland Type",     f"CCG {r.gland_type}")
        data_row_2col("Glands End A",   str(r.total_glands_a))
        data_row_2col("Glands End B",   str(r.total_glands_b))
        data_row_2col("Total Glands",   str(r.total_glands_a + r.total_glands_b))
        row += 1

        # Per-cable selection table
        section_hdr("Cable-by-Cable Gland Selection")
        for ci, hd in enumerate(["Cable No.", "End A", "End B"], 1):
            c = ws.cell(row, ci, hd)
            c.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
            c.fill = fill(NAVY)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = bdr
        ws.row_dimensions[row].height = 20
        row += 1
        for i, (a, b) in enumerate(r.selections):
            bg_a = GREEN if a else "F4F6F8"
            bg_b = GREEN if b else "F4F6F8"
            tc_a = WHITE if a else "000000"
            tc_b = WHITE if b else "000000"
            ws.cell(row, 1, str(i+1)).border = bdr
            ca = ws.cell(row, 2, "Gland" if a else "—")
            ca.fill = fill(bg_a); ca.font = Font(name="Calibri", size=10, color=tc_a, bold=a)
            ca.alignment = Alignment(horizontal="center"); ca.border = bdr
            cb = ws.cell(row, 3, "Gland" if b else "—")
            cb.fill = fill(bg_b); cb.font = Font(name="Calibri", size=10, color=tc_b, bold=b)
            cb.alignment = Alignment(horizontal="center"); cb.border = bdr
            ws.row_dimensions[row].height = 18
            row += 1
        row += 1

        # Gland recommendation
        section_hdr("Recommended Gland")
        m = r.gland_match
        if m and m.matched:
            data_row_2col("Product Code",   m.product_code)
            data_row_2col("Size Reference", m.size_ref)
            data_row_2col("Thread",         m.thread)
            data_row_2col("OD Range",       f"{m.od_min:.1f} – {m.od_max:.1f} mm")
            data_row_2col("Cable OD",       f"{m.cable_od:.1f} mm")
            ws.merge_cells(f"A{row}:E{row}")
            c = ws.cell(row, 1, "PASS — Gland fits cable OD")
            c.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
            c.fill = fill(GREEN)
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws.row_dimensions[row].height = 18
            row += 1
        row += 1

        # BOQ
        if r.boq:
            section_hdr("Bill of Quantities")
            boq_hdrs = ["No.", "Description", "Product Code", "Qty", "Unit"]
            for ci, hd in enumerate(boq_hdrs, 1):
                c = ws.cell(row, ci, hd)
                c.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
                c.fill = fill(NAVY)
                c.alignment = Alignment(
                    horizontal="center" if ci != 2 else "left",
                    vertical="center", indent=1 if ci == 2 else 0)
                c.border = bdr
            ws.row_dimensions[row].height = 22
            row += 1
            for i, item in enumerate(r.boq):
                bg = LIGHT if i % 2 == 0 else WHITE
                for ci, val in enumerate(
                    [str(item.item_no), item.description,
                     item.product_code, str(item.qty), item.unit], 1
                ):
                    c = ws.cell(row, ci, val)
                    c.font = Font(name="Calibri", size=10)
                    c.fill = fill(bg)
                    c.alignment = Alignment(
                        horizontal="center" if ci in (1, 4, 5) else "left",
                        vertical="center", indent=1 if ci == 2 else 0)
                    c.border = bdr
                ws.row_dimensions[row].height = 18
                row += 1
            row += 2

        ws.merge_cells(f"A{row}:E{row}")
        c = ws.cell(row, 1,
            f"Engineering Calculator Suite  •  Solareff Engineering  •  {now}  •  "
            "CCG GI010214 / GI020816  •  Verify OD against actual cable before ordering.")
        c.font = Font(name="Calibri", size=8, italic=True, color="AAAAAA")
        c.alignment = Alignment(horizontal="center")
        wb.save(filepath)

    # ─────────────────────────────────────────── BaseCalculator API ────────

    def get_inputs(self) -> dict:
        d = {
            "Cable Type":  self.cable_type_combo.currentData() or "",
            "Cable Size":  f"{self.cable_size_combo.currentData() or 0:.0f} mm²",
            "Quantity":    str(self.cable_qty_spin.value()),
        }
        if self.run_label_edit.text().strip():
            d["Run Label"] = self.run_label_edit.text().strip()
        if self.lug_enable_chk.isChecked():
            d["Lug Size"]  = f"{self.lug_size_combo.currentData() or 0:.0f} mm²"
            d["Lug Bolt"]  = self.lug_bolt_combo.currentData() or ""
            d["Lug Cores"] = str(self.lug_cores_spin.value())
        return d

    def get_results(self) -> dict:
        if not self._result:
            return {}
        r = self._result
        d: dict = {
            "Cable OD":       f"{r.cable_od:.1f} mm" if r.cable_od > 0 else "—",
            "Gland Type":     f"CCG {r.gland_type}",
            "Glands End A":   str(r.total_glands_a),
            "Glands End B":   str(r.total_glands_b),
            "Total Glands":   str(r.total_glands_a + r.total_glands_b),
        }
        m = r.gland_match
        if m and m.matched:
            d["Recommended Gland"] = (
                f"{m.product_code}  |  {m.size_ref}  |  {m.thread}  |  "
                f"OD {m.od_min:.1f}–{m.od_max:.1f}mm  (PASS)")
        elif r.total_glands_a + r.total_glands_b > 0:
            d["Recommended Gland"] = "NO MATCH"
        for item in r.boq:
            d[f"BOQ {item.item_no}"] = f"{item.description}  ×{item.qty}"
        return d

    def reset(self):
        self.cable_type_combo.setCurrentIndex(0)
        self.cable_qty_spin.setValue(1)
        self.run_label_edit.clear()
        self.lug_enable_chk.setChecked(False)
        self.lug_cores_spin.setValue(3)
        self.lug_bolt_combo.setCurrentIndex(3)
        self._result = None
        self.gland_card.set_result(GlandCalcResult(
            "", 0, 0, 0, "", [], None, 0, 0, []))
        self.boq_table.setRowCount(0)
        self._on_cable_type_changed()
        self._on_qty_changed(1)
