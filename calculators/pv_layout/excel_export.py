"""
excel_export.py
---------------
Export PV stringing layout to an Excel workbook (.xlsx).

Sheet 1  "PV Layout"        — scaled grid, one cell per panel, coloured by string
Sheet 2  "String Schedule"  — table: string name, inverter, MPPT, panel count, section, angle
Sheet 3  "Project Info"     — metadata

Rotation fix
------------
DXF buildings are often drawn at an angle (e.g. 30° from horizontal).  Mapping
raw DXF (cx, cy) → Excel (col, row) directly produces a staircase/diagonal
pattern because every step along a physical panel row also shifts the column.

Fix: detect the dominant panel rotation angle, de-rotate all panel centroids
around the cloud centre by −angle before computing Excel col/row.  This aligns
the physical panel rows with Excel rows, producing a clean rectangular grid.
"""
from __future__ import annotations

import datetime
import math
from collections import Counter
from typing import List

from calculators.pv_layout.canvas_view import PANEL_COLORS


# ─────────────────────────────────────────────────────────────────────────────
# Colour helpers
# ─────────────────────────────────────────────────────────────────────────────

def _to_argb(hex_color: str) -> str:
    """'#rrggbb' → 'FFrrggbb' (openpyxl ARGB, fully opaque)."""
    return "FF" + hex_color.lstrip("#").upper()


# ─────────────────────────────────────────────────────────────────────────────
# Rotation helpers
# ─────────────────────────────────────────────────────────────────────────────

def _dominant_rotation(panels) -> float:
    """
    Return the most common panel rotation angle (degrees).

    Panels on the same roof face share a rotation; this returns the angle that
    covers the most panels.  Result is in [0, 360).
    """
    if not panels:
        return 0.0
    # Round to nearest 1° so small floating-point noise doesn't split the mode
    rots = [round(p.rotation % 360) for p in panels]
    return float(Counter(rots).most_common(1)[0][0])


def _derotate(panels, angle_deg: float) -> dict:
    """
    Rotate every panel centroid by *−angle_deg* around the cloud centroid.

    Returns ``{panel_id: (rotated_cx, rotated_cy)}``.

    Panels that are already axis-aligned (angle ≈ 0° or 180°) are returned
    unchanged so we don't introduce floating-point drift.
    """
    if not panels:
        return {}

    # Skip de-rotation if already close to axis-aligned
    norm = angle_deg % 180
    if norm < 1.5 or norm > 178.5:
        return {p.id: (p.cx, p.cy) for p in panels}

    mean_x = sum(p.cx for p in panels) / len(panels)
    mean_y = sum(p.cy for p in panels) / len(panels)
    rad    = math.radians(-angle_deg)   # negative = undo the rotation
    cos_a, sin_a = math.cos(rad), math.sin(rad)

    result = {}
    for p in panels:
        dx, dy = p.cx - mean_x, p.cy - mean_y
        result[p.id] = (
            dx * cos_a - dy * sin_a + mean_x,
            dx * sin_a + dy * cos_a + mean_y,
        )
    return result


# ─────────────────────────────────────────────────────────────────────────────
# Public entry point
# ─────────────────────────────────────────────────────────────────────────────

def export_layout(strings, panels, roof_sections, filepath: str, metadata: dict):
    """
    Parameters
    ----------
    strings       : list[StringResult]
    panels        : list[ParsedPanel]
    roof_sections : list[dict]  — each has keys id, name, angle (and optionally azimuth)
    filepath      : str  — destination .xlsx path
    metadata      : dict — keys: project, client, location, prepared_by, date
    """
    from openpyxl import Workbook

    wb = Workbook()

    # Build lookup: panel_id → (StringResult, order_index)
    pid_to_str: dict = {}
    for sr in strings:
        for i, pid in enumerate(sr.panels):
            pid_to_str[pid] = (sr, i)

    _sheet_layout(wb.active, panels, pid_to_str, metadata)
    _sheet_schedule(wb.create_sheet("String Schedule"), strings, roof_sections)
    _sheet_project(wb.create_sheet("Project Info"), panels, strings, pid_to_str, metadata)

    wb.save(filepath)


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 1 — PV Layout grid
# ─────────────────────────────────────────────────────────────────────────────

def _sheet_layout(ws, panels, pid_to_str, metadata):
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    ws.title = "PV Layout"
    ws.sheet_view.showGridLines = False

    thin = Side(style="thin", color="30363D")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── title banner ──────────────────────────────────────────────────────────
    ws.merge_cells("A1:J1")
    t = ws.cell(1, 1, "PV LAYOUT  —  STRING DESIGN")
    t.font      = Font(name="Segoe UI", size=13, bold=True, color="58A6FF")
    t.fill      = PatternFill("solid", fgColor="FF1F3A5F")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:J2")
    sub_txt = (
        f"Project: {metadata.get('project','—')}   |   "
        f"Client: {metadata.get('client','—')}   |   "
        f"Prepared by: {metadata.get('prepared_by','—')}   |   "
        f"Date: {metadata.get('date', datetime.date.today().strftime('%d %B %Y'))}"
    )
    s = ws.cell(2, 1, sub_txt)
    s.font      = Font(name="Segoe UI", size=9, color="8B949E")
    s.fill      = PatternFill("solid", fgColor="FF161B22")
    s.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16

    if not panels:
        return

    # ── de-rotate coordinates so physical rows align with Excel rows ──────────
    rot_angle = _dominant_rotation(panels)
    derot     = _derotate(panels, rot_angle)   # {id: (rx, ry)}

    # Panel average dimensions (used as grid step)
    avg_w = sum(p.width  for p in panels) / len(panels)
    avg_h = sum(p.height for p in panels) / len(panels)

    # After de-rotation the panel's width aligns with X and height with Y.
    # Use a 1.15× factor (snug grid) so cells sit close together.
    step_x = max(avg_w, 0.01) * 1.15
    step_y = max(avg_h, 0.01) * 1.15

    rx_vals = [derot[p.id][0] for p in panels]
    ry_vals = [derot[p.id][1] for p in panels]
    min_rx  = min(rx_vals)
    max_ry  = max(ry_vals)   # DXF Y grows upward → flip when mapping to rows

    ROW_OFF = 4    # first data row  (rows 1–3 are header)
    COL_OFF = 2    # first data column

    COL_W = 5.5    # Excel column width
    ROW_H = 22.0   # Excel row height (points)

    panel_cells: dict = {}   # panel_id → (excel_row, excel_col)
    max_c, max_r = COL_OFF, ROW_OFF

    for p in panels:
        rx, ry = derot[p.id]
        ec = COL_OFF + round((rx  - min_rx) / step_x)
        er = ROW_OFF + round((max_ry - ry)  / step_y)   # flip Y → rows
        panel_cells[p.id] = (er, ec)
        max_c = max(max_c, ec)
        max_r = max(max_r, er)

    # ── set uniform column widths and row heights ─────────────────────────────
    for c in range(1, max_c + 4):
        ws.column_dimensions[get_column_letter(c)].width = COL_W
    for r in range(1, max_r + 4):
        ws.row_dimensions[r].height = ROW_H

    # ── fill panel cells ──────────────────────────────────────────────────────
    for p in panels:
        er, ec = panel_cells[p.id]
        cell   = ws.cell(row=er, column=ec)

        if p.id in pid_to_str:
            sr, order  = pid_to_str[p.id]
            color_hex  = PANEL_COLORS[sr.color_index % len(PANEL_COLORS)]
            is_start   = (order == 0)
            is_end     = (order == len(sr.panels) - 1)

            if is_start:
                label      = f"+ {sr.name}"
                font_color = "3FB950"
                bold       = True
            elif is_end:
                label      = f"- {sr.name}"
                font_color = "F85149"
                bold       = True
            else:
                label      = sr.name
                font_color = "FFFFFF"
                bold       = False

            cell.value = label
            cell.fill  = PatternFill("solid", fgColor=_to_argb(color_hex))
            cell.font  = Font(name="Segoe UI", size=7, bold=bold, color=font_color)
        else:
            cell.value = "—"
            cell.fill  = PatternFill("solid", fgColor="FF1A1F27")
            cell.font  = Font(name="Segoe UI", size=7, color="484F58")

        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   shrink_to_fit=True)
        cell.border = bdr

    # ── rotation note in header ───────────────────────────────────────────────
    if abs(rot_angle % 180) > 1.5:
        note_row = 3
        ws.merge_cells(f"A{note_row}:J{note_row}")
        note = ws.cell(note_row, 1,
            f"Note: Panel array de-rotated {rot_angle:.0f}° to align with grid  "
            f"(dominant panel rotation in DXF)")
        note.font      = Font(name="Segoe UI", size=8, italic=True, color="8B949E")
        note.fill      = PatternFill("solid", fgColor="FF161B22")
        note.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[note_row].height = 14

    # ── legend block ──────────────────────────────────────────────────────────
    legend_col = max_c + 2
    ws.cell(ROW_OFF, legend_col, "LEGEND").font = Font(
        name="Segoe UI", size=9, bold=True, color="E6EDF3")
    ws.cell(ROW_OFF + 1, legend_col, "+ panel").font = Font(
        name="Segoe UI", size=8, bold=True, color="3FB950")
    ws.cell(ROW_OFF + 2, legend_col, "- panel").font = Font(
        name="Segoe UI", size=8, bold=True, color="F85149")
    ws.cell(ROW_OFF + 3, legend_col, "unassigned").fill = PatternFill(
        "solid", fgColor="FF1A1F27")
    ws.cell(ROW_OFF + 3, legend_col).font = Font(
        name="Segoe UI", size=8, color="484F58")
    for c in [legend_col, legend_col + 1]:
        ws.column_dimensions[get_column_letter(c)].width = 14

    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize   = 9    # A4
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 2 — String Schedule
# ─────────────────────────────────────────────────────────────────────────────

def _sheet_schedule(ws, strings, roof_sections):
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    ws.sheet_view.showGridLines = False

    headers    = ["String", "Inverter", "MPPT", "No. Panels", "Roof Section", "Angle (°)", "Azimuth (°)"]
    col_widths = [14,       10,          8,      12,            18,             10,           12]

    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
        c = ws.cell(1, i, h)
        c.font      = Font(name="Segoe UI", size=10, bold=True, color="E6EDF3")
        c.fill      = PatternFill("solid", fgColor="FF1F3A5F")
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    sec_map = {rs["id"]: rs for rs in roof_sections} if roof_sections else {}
    thin    = Side(style="thin", color="30363D")
    bdr     = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_i, sr in enumerate(strings, 2):
        sec       = sec_map.get(sr.roof_section_id, {})
        color_hex = PANEL_COLORS[sr.color_index % len(PANEL_COLORS)]
        row_vals  = [
            sr.name,
            sr.inverter_id,
            sr.mppt_letter,
            len(sr.panels),
            sec.get("name",    "—"),
            sec.get("angle",   "—"),
            sec.get("azimuth", "—"),
        ]
        for col_i, val in enumerate(row_vals, 1):
            c = ws.cell(row_i, col_i, val)
            c.fill      = PatternFill("solid",
                fgColor=_to_argb(color_hex) if col_i == 1 else "FF0D1117")
            c.font      = Font(name="Segoe UI", size=9,
                               color="FFFFFF" if col_i == 1 else "C9D1D9")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = bdr
        ws.row_dimensions[row_i].height = 18


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 3 — Project Info
# ─────────────────────────────────────────────────────────────────────────────

def _sheet_project(ws, panels, strings, pid_to_str, metadata):
    from openpyxl.styles import PatternFill, Font, Alignment

    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 44

    ws.merge_cells("A1:B1")
    t = ws.cell(1, 1, "PROJECT INFORMATION")
    t.font      = Font(name="Segoe UI", size=12, bold=True, color="58A6FF")
    t.fill      = PatternFill("solid", fgColor="FF1F3A5F")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    unassigned = sum(1 for p in panels if p.id not in pid_to_str)
    rows = [
        ("Project",         metadata.get("project",     "")),
        ("Client",          metadata.get("client",      "")),
        ("Location",        metadata.get("location",    "")),
        ("Prepared by",     metadata.get("prepared_by", "")),
        ("Date",            metadata.get("date", datetime.date.today().strftime("%d %B %Y"))),
        ("",                ""),
        ("Total panels",    str(len(panels))),
        ("Total strings",   str(len(strings))),
        ("Assigned panels", str(len(panels) - unassigned)),
        ("Unassigned",      str(unassigned)),
    ]

    for r, (lbl, val) in enumerate(rows, 2):
        l = ws.cell(r, 1, lbl)
        l.font = Font(name="Segoe UI", size=10, bold=bool(lbl), color="8B949E")
        l.fill = PatternFill("solid", fgColor="FF161B22")
        v = ws.cell(r, 2, val)
        v.font = Font(name="Segoe UI", size=10, color="E6EDF3")
        v.fill = PatternFill("solid", fgColor="FF0D1117")
        ws.row_dimensions[r].height = 18
