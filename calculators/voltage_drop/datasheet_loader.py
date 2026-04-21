"""
datasheet_loader.py
-------------------
Reads cable datasheet files (Excel .xlsx or CSV) and returns structured cable data.
Also generates the blank Excel template for users to fill in.

Expected columns (case-insensitive):
    cable_name             - Human-readable cable description
    size_mm2               - Conductor cross-section in mm²
    resistance_ohm_per_km  - DC/AC resistance in Ω/km (single conductor)
    mv_per_am              - (optional) mV/A/m from manufacturer (single conductor)
    notes                  - (optional) Free-text notes
"""

import os
import csv
from typing import Optional


REQUIRED_COLUMNS = {"cable_name", "size_mm2", "resistance_ohm_per_km"}
OPTIONAL_COLUMNS = {"mv_per_am", "notes"}
ALL_COLUMNS = REQUIRED_COLUMNS | OPTIONAL_COLUMNS

TEMPLATE_HEADERS = [
    "cable_name",
    "size_mm2",
    "resistance_ohm_per_km",
    "mv_per_am",
    "notes",
]

TEMPLATE_EXAMPLE_ROWS = [
    ["16mm² CU XLPE 0.6/1kV",  16.0,   1.150,  1.91,  "Aberdare Cat. Ref A123"],
    ["25mm² CU XLPE 0.6/1kV",  25.0,   0.727,  1.20,  "Aberdare Cat. Ref A124"],
    ["35mm² CU XLPE 0.6/1kV",  35.0,   0.524,  0.87,  "Aberdare Cat. Ref A125"],
    ["50mm² AL XLPE 0.6/1kV",  50.0,   0.641,  1.11,  "Nexans AL cable"],
    ["70mm² AL XLPE 0.6/1kV",  70.0,   0.443,  0.77,  "Nexans AL cable"],
]


def generate_template(filepath: str) -> None:
    """
    Write an Excel template file that users can fill in from manufacturer datasheets.
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Cable Data"

    NAVY  = "1F3F60"
    WHITE = "FFFFFF"
    GREY  = "F4F6F8"
    YELLOW = "FFF9C4"

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Column widths
    widths = [32, 12, 22, 14, 40]
    for i, w in enumerate(widths, 1):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(i)].width = w

    # Title row
    ws.merge_cells("A1:E1")
    c = ws.cell(1, 1, "Engineering Suite — Cable Datasheet Template")
    c.font = Font(name="Calibri", bold=True, size=13, color=WHITE)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Instruction row
    ws.merge_cells("A2:E2")
    c = ws.cell(2, 1,
        "Fill in your cables below. Required columns: cable_name, size_mm2, resistance_ohm_per_km. "
        "mv_per_am and notes are optional. Do NOT rename or remove column headers.")
    c.font = Font(name="Calibri", size=9, italic=True, color="555555")
    c.fill = PatternFill("solid", fgColor=YELLOW)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    ws.row_dimensions[2].height = 32

    # Header row
    for col_i, header in enumerate(TEMPLATE_HEADERS, 1):
        c = ws.cell(3, col_i, header)
        c.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border = border
    ws.row_dimensions[3].height = 22

    # Example rows
    for row_i, row_data in enumerate(TEMPLATE_EXAMPLE_ROWS, 4):
        bg = GREY if row_i % 2 == 0 else WHITE
        for col_i, val in enumerate(row_data, 1):
            c = ws.cell(row_i, col_i, val)
            c.font = Font(name="Calibri", size=10)
            c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            c.border = border
        ws.row_dimensions[row_i].height = 18

    wb.save(filepath)


def load_datasheet(filepath: str) -> list[dict]:
    """
    Load a cable datasheet from an Excel (.xlsx) or CSV (.csv) file.

    Returns:
        List of cable dicts with keys:
            cable_name, size_mm2, resistance_ohm_per_km, mv_per_am (or None), notes (or "")

    Raises:
        ValueError if required columns are missing or file cannot be parsed.
    """
    ext = os.path.splitext(filepath)[1].lower()
    if ext == ".xlsx":
        return _load_excel(filepath)
    elif ext == ".csv":
        return _load_csv(filepath)
    else:
        raise ValueError(f"Unsupported file type: {ext}. Please use .xlsx or .csv")


def _normalise_headers(headers: list[str]) -> dict[str, int]:
    """Map normalised column name → column index."""
    return {h.strip().lower().replace(" ", "_"): i for i, h in enumerate(headers)}


def _validate_columns(header_map: dict) -> None:
    missing = REQUIRED_COLUMNS - set(header_map.keys())
    if missing:
        raise ValueError(
            f"Missing required columns: {', '.join(sorted(missing))}\n"
            f"Required: {', '.join(sorted(REQUIRED_COLUMNS))}"
        )


def _parse_float(val) -> Optional[float]:
    if val is None or str(val).strip() in ("", "-", "N/A", "n/a"):
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _load_excel(filepath: str) -> list[dict]:
    from openpyxl import load_workbook
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("The Excel file appears to be empty.")

    # Find the header row — it must contain 'cable_name'
    header_row_idx = None
    for i, row in enumerate(rows):
        row_lower = [str(c).strip().lower() if c else "" for c in row]
        if "cable_name" in row_lower:
            header_row_idx = i
            break

    if header_row_idx is None:
        raise ValueError(
            "Could not find the header row. "
            "Make sure your file has a row with 'cable_name' as a column header."
        )

    headers = [str(c).strip().lower().replace(" ", "_") if c else "" for c in rows[header_row_idx]]
    header_map = {h: i for i, h in enumerate(headers) if h}
    _validate_columns(header_map)

    cables = []
    for row in rows[header_row_idx + 1:]:
        if not any(row):
            continue  # skip empty rows
        name = str(row[header_map["cable_name"]]).strip() if row[header_map["cable_name"]] else ""
        if not name or name.lower() in ("", "none"):
            continue

        r_val = _parse_float(row[header_map["resistance_ohm_per_km"]])
        if r_val is None:
            continue  # skip rows without resistance value

        mv_val = _parse_float(row[header_map["mv_per_am"]]) if "mv_per_am" in header_map else None
        size_val = _parse_float(row[header_map["size_mm2"]])
        notes_idx = header_map.get("notes")
        notes = str(row[notes_idx]).strip() if notes_idx is not None and row[notes_idx] else ""

        cables.append({
            "cable_name":            name,
            "size_mm2":              size_val,
            "resistance_ohm_per_km": r_val,
            "mv_per_am":             mv_val,
            "notes":                 notes,
        })

    wb.close()
    if not cables:
        raise ValueError("No valid cable entries found in the file. Check that data rows are below the header row.")

    return cables


def _load_csv(filepath: str) -> list[dict]:
    cables = []
    with open(filepath, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        if reader.fieldnames is None:
            raise ValueError("CSV file has no headers.")

        header_map = _normalise_headers(list(reader.fieldnames))
        _validate_columns(header_map)

        for row in reader:
            norm = {k.strip().lower().replace(" ", "_"): v for k, v in row.items()}
            name = norm.get("cable_name", "").strip()
            if not name:
                continue
            r_val = _parse_float(norm.get("resistance_ohm_per_km"))
            if r_val is None:
                continue
            cables.append({
                "cable_name":            name,
                "size_mm2":              _parse_float(norm.get("size_mm2")),
                "resistance_ohm_per_km": r_val,
                "mv_per_am":             _parse_float(norm.get("mv_per_am")),
                "notes":                 norm.get("notes", "").strip(),
            })

    if not cables:
        raise ValueError("No valid cable entries found in the CSV file.")
    return cables
