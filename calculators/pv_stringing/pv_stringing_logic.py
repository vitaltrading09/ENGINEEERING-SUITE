"""
pv_stringing_logic.py
---------------------
PV String Sizing calculations.

Temperature correction follows IEC 61215 / IEC 62548:
  V_oc(T) = V_oc_stc × [1 + (β_Voc / 100) × (T - 25)]
  V_mp(T) = V_mp_stc × [1 + (β_Vmp / 100) × (T - 25)]
  I_sc(T) = I_sc_stc × [1 + (α_Isc / 100) × (T - 25)]

String sizing rules:
  Max series panels  = floor(V_inverter_max  / V_oc_cold)   — open-circuit at min temp
  Min series panels  = ceil (V_mppt_min      / V_mp_hot)    — MPPT lower edge at max temp
  Max series (MPPT)  = floor(V_mppt_max      / V_mp_hot)    — MPPT upper edge at max temp
  Recommended series = largest value in [min_series .. min(max_series, max_mppt_series)]
  Max parallel strings per MPPT = floor(I_mppt_max / I_sc_hot)
"""

import math
from dataclasses import dataclass
from typing import Optional


@dataclass
class StringResult:
    # Temperature-corrected values
    voc_cold: float          # Voc at T_min  (V)
    vmp_hot:  float          # Vmp at T_max  (V)
    isc_hot:  float          # Isc at T_max  (A)

    # Series string limits
    max_series_voc:   int    # floor(V_inv_max  / voc_cold)
    min_series_mppt:  int    # ceil (V_mppt_min / vmp_hot)
    max_series_mppt:  int    # floor(V_mppt_max / vmp_hot)
    recommended_series: int  # best fit inside MPPT window

    # Parallel limits
    max_parallel: int        # floor(I_mppt_max / isc_hot)

    # Per-string & array totals (using recommended_series)
    voc_string:   float      # recommended_series × voc_cold
    vmp_string:   float      # recommended_series × vmp_hot
    power_stc_string: float  # recommended_series × P_max_stc  (W)
    power_stc_array:  float  # recommended_series × max_parallel × P_max_stc  (W)

    # Validity
    valid: bool
    warnings: list


def calc_string_sizing(
    # Panel STC values
    voc_stc:   float,   # V
    vmp_stc:   float,   # V
    isc_stc:   float,   # A
    pmax_stc:  float,   # W
    # Temperature coefficients (%/°C)
    beta_voc:  float,   # typically negative, e.g. -0.30
    beta_vmp:  float,   # typically negative, e.g. -0.35
    alpha_isc: float,   # typically positive, e.g. +0.05
    # Site temperatures
    t_min: float,       # °C  (lowest ambient, e.g. -5)
    t_max: float,       # °C  (highest cell temp, e.g. 70)
    # Inverter / MPPT limits
    v_inv_max:  float,  # V   absolute max DC input voltage
    v_mppt_min: float,  # V   MPPT window lower edge
    v_mppt_max: float,  # V   MPPT window upper edge
    i_mppt_max: float,  # A   max input current per MPPT tracker
) -> StringResult:

    warnings = []

    # Temperature-corrected values
    voc_cold = voc_stc * (1.0 + (beta_voc  / 100.0) * (t_min - 25.0))
    vmp_hot  = vmp_stc * (1.0 + (beta_vmp  / 100.0) * (t_max - 25.0))
    isc_hot  = isc_stc * (1.0 + (alpha_isc / 100.0) * (t_max - 25.0))

    # Guard against zero/negative corrected values
    voc_cold = max(voc_cold, 0.01)
    vmp_hot  = max(vmp_hot,  0.01)
    isc_hot  = max(isc_hot,  0.01)

    # Series limits
    max_series_voc  = int(math.floor(v_inv_max  / voc_cold))
    min_series_mppt = int(math.ceil (v_mppt_min / vmp_hot))
    max_series_mppt = int(math.floor(v_mppt_max / vmp_hot))

    # Effective max series (must satisfy both Voc and MPPT upper limit)
    effective_max = min(max_series_voc, max_series_mppt)

    valid = True
    if effective_max < min_series_mppt:
        warnings.append(
            "No valid series count satisfies both the inverter voltage limit "
            "and the MPPT window. Check panel and inverter specifications."
        )
        valid = False
        recommended_series = min_series_mppt  # best guess
    else:
        recommended_series = effective_max    # maximise energy yield

    if recommended_series < 1:
        recommended_series = 1
        valid = False
        warnings.append("Recommended series count fell below 1 — check inputs.")

    # Parallel limit
    max_parallel = int(math.floor(i_mppt_max / isc_hot))
    if max_parallel < 1:
        max_parallel = 1
        warnings.append(
            "Isc (hot) exceeds inverter MPPT max current — single string only; "
            "verify inverter spec."
        )

    # Derived totals
    voc_string        = recommended_series * voc_cold
    vmp_string        = recommended_series * vmp_hot
    power_stc_string  = recommended_series * pmax_stc
    power_stc_array   = power_stc_string   * max_parallel

    return StringResult(
        voc_cold=round(voc_cold, 3),
        vmp_hot =round(vmp_hot,  3),
        isc_hot =round(isc_hot,  3),
        max_series_voc   =max_series_voc,
        min_series_mppt  =min_series_mppt,
        max_series_mppt  =max_series_mppt,
        recommended_series=recommended_series,
        max_parallel     =max_parallel,
        voc_string       =round(voc_string, 2),
        vmp_string       =round(vmp_string, 2),
        power_stc_string =round(power_stc_string, 1),
        power_stc_array  =round(power_stc_array,  1),
        valid   =valid,
        warnings=warnings,
    )


# ─────────────────────────────────────────────────────────────────────────────
# Excel Export (openpyxl)
# ─────────────────────────────────────────────────────────────────────────────

def export_to_excel(
    result: StringResult,
    filepath: str,
    # metadata
    project:    str = "",
    client:     str = "",
    location:   str = "",
    prepared_by:str = "",
    # raw inputs (for the inputs sheet)
    inputs: Optional[dict] = None,
):
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── colour palette ────────────────────────────────────────────────────────
    C_HEADER_BG  = "1F3A5F"   # dark blue
    C_HEADER_FG  = "FFFFFF"
    C_ACCENT     = "58A6FF"   # blue accent
    C_PASS       = "238636"   # green
    C_WARN       = "D29922"   # amber
    C_FAIL       = "DA3633"   # red
    C_LABEL_BG   = "161B22"
    C_LABEL_FG   = "8B949E"
    C_VALUE_FG   = "E6EDF3"
    C_ALT_BG     = "0D1117"
    C_BORDER     = "30363D"

    thin = Side(style="thin", color=C_BORDER)
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr_font(sz=11, bold=True):
        return Font(name="Segoe UI", size=sz, bold=bold, color=C_HEADER_FG)

    def lbl_font(sz=10, bold=False):
        return Font(name="Segoe UI", size=sz, bold=bold, color=C_LABEL_FG)

    def val_font(sz=10, bold=False):
        return Font(name="Segoe UI", size=sz, bold=bold, color=C_VALUE_FG)

    def accent_font(sz=11, bold=True):
        return Font(name="Segoe UI", size=sz, bold=bold, color=C_ACCENT)

    # ── Sheet 1 : Summary ─────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "String Sizing Summary"
    ws.sheet_view.showGridLines = False

    # column widths
    for col, w in [(1, 5), (2, 36), (3, 22), (4, 22), (5, 5)]:
        ws.column_dimensions[get_column_letter(col)].width = w

    row = 1

    # Title banner
    ws.merge_cells(f"B{row}:D{row}")
    c = ws.cell(row, 2, "PV STRING SIZING REPORT")
    c.font      = Font(name="Segoe UI", size=16, bold=True, color=C_ACCENT)
    c.fill      = PatternFill("solid", fgColor=C_HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 36
    for col in [3, 4]:
        ws.cell(row, col).fill = PatternFill("solid", fgColor=C_HEADER_BG)
    row += 1

    # Meta rows
    import datetime
    meta = [
        ("Project",     project),
        ("Client",      client),
        ("Location",    location),
        ("Prepared by", prepared_by),
        ("Date",        datetime.date.today().strftime("%d %B %Y")),
    ]
    for lbl, val in meta:
        l = ws.cell(row, 2, lbl)
        l.font = lbl_font(bold=True); l.fill = PatternFill("solid", fgColor=C_LABEL_BG)
        l.alignment = Alignment(vertical="center", indent=1)
        ws.merge_cells(f"C{row}:D{row}")
        v = ws.cell(row, 3, val)
        v.font = val_font(); v.fill = PatternFill("solid", fgColor=C_ALT_BG)
        v.alignment = Alignment(vertical="center", indent=1)
        ws.row_dimensions[row].height = 18
        row += 1

    row += 1  # spacer

    def section_title(r, title):
        ws.merge_cells(f"B{r}:D{r}")
        c = ws.cell(r, 2, f"  {title}")
        c.font = Font(name="Segoe UI", size=10, bold=True, color=C_HEADER_FG)
        c.fill = PatternFill("solid", fgColor=C_HEADER_BG)
        c.alignment = Alignment(vertical="center")
        ws.row_dimensions[r].height = 20
        return r + 1

    def data_row(r, label, value, unit="", highlight=None):
        l = ws.cell(r, 2, label)
        l.font = lbl_font()
        l.fill = PatternFill("solid", fgColor=C_LABEL_BG)
        l.alignment = Alignment(vertical="center", indent=2)
        l.border = bdr

        display = f"{value}  {unit}".strip() if unit else str(value)
        v = ws.cell(r, 3, display)
        if highlight == "pass":
            v.font = Font(name="Segoe UI", size=10, bold=True, color=C_PASS)
        elif highlight == "warn":
            v.font = Font(name="Segoe UI", size=10, bold=True, color=C_WARN)
        elif highlight == "fail":
            v.font = Font(name="Segoe UI", size=10, bold=True, color=C_FAIL)
        elif highlight == "accent":
            v.font = accent_font(sz=10)
        else:
            v.font = val_font()
        v.fill      = PatternFill("solid", fgColor=C_ALT_BG)
        v.alignment = Alignment(vertical="center", indent=1)
        v.border    = bdr

        ws.merge_cells(f"C{r}:D{r}")
        ws.row_dimensions[r].height = 18

    # ── Temperature Corrections ──
    row = section_title(row, "TEMPERATURE-CORRECTED VALUES")
    data_row(row, "Voc at minimum temperature (Voc_cold)", f"{result.voc_cold:.3f}", "V"); row += 1
    data_row(row, "Vmp at maximum temperature (Vmp_hot)",  f"{result.vmp_hot:.3f}",  "V"); row += 1
    data_row(row, "Isc at maximum temperature (Isc_hot)",  f"{result.isc_hot:.3f}",  "A"); row += 1
    row += 1

    # ── Series String Sizing ──
    row = section_title(row, "SERIES STRING SIZING")
    data_row(row, "Max panels in series (Voc limit)",    result.max_series_voc,   "panels"); row += 1
    data_row(row, "Min panels in series (MPPT lower)",   result.min_series_mppt,  "panels"); row += 1
    data_row(row, "Max panels in series (MPPT upper)",   result.max_series_mppt,  "panels"); row += 1

    hl = "pass" if result.valid else "warn"
    data_row(row, "⭐  RECOMMENDED series count",        result.recommended_series, "panels", highlight=hl); row += 1
    row += 1

    # ── Parallel String Sizing ──
    row = section_title(row, "PARALLEL STRING SIZING")
    data_row(row, "Max parallel strings per MPPT", result.max_parallel, "strings"); row += 1
    row += 1

    # ── Array Totals ──
    row = section_title(row, "ARRAY TOTALS  (recommended series × max parallel)")
    data_row(row, "Voc per string (cold)",      f"{result.voc_string:.2f}",      "V", "accent"); row += 1
    data_row(row, "Vmp per string (hot)",       f"{result.vmp_string:.2f}",      "V", "accent"); row += 1
    data_row(row, "Power per string (STC)",     f"{result.power_stc_string:.0f}","W", "accent"); row += 1
    data_row(row, "Total array power (STC)",    f"{result.power_stc_array/1000:.2f}", "kWp", "accent"); row += 1
    row += 1

    # ── Warnings ──
    if result.warnings:
        row = section_title(row, "WARNINGS")
        for w in result.warnings:
            ws.merge_cells(f"B{row}:D{row}")
            c = ws.cell(row, 2, f"  ⚠  {w}")
            c.font = Font(name="Segoe UI", size=10, color=C_WARN)
            c.fill = PatternFill("solid", fgColor="2D2000")
            c.alignment = Alignment(vertical="center", wrap_text=True)
            ws.row_dimensions[row].height = 30
            row += 1

    # ── Sheet 2 : Inputs ──────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Inputs")
    ws2.sheet_view.showGridLines = False
    for col, w in [(1, 5), (2, 36), (3, 22), (4, 5)]:
        ws2.column_dimensions[get_column_letter(col)].width = w

    r2 = 1
    ws2.merge_cells(f"B{r2}:C{r2}")
    c = ws2.cell(r2, 2, "INPUT PARAMETERS")
    c.font = Font(name="Segoe UI", size=13, bold=True, color=C_ACCENT)
    c.fill = PatternFill("solid", fgColor=C_HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[r2].height = 30
    ws2.cell(r2, 3).fill = PatternFill("solid", fgColor=C_HEADER_BG)
    r2 += 2

    if inputs:
        for section, params in inputs.items():
            # section heading
            ws2.merge_cells(f"B{r2}:C{r2}")
            sh = ws2.cell(r2, 2, f"  {section}")
            sh.font = Font(name="Segoe UI", size=10, bold=True, color=C_HEADER_FG)
            sh.fill = PatternFill("solid", fgColor=C_HEADER_BG)
            sh.alignment = Alignment(vertical="center")
            ws2.row_dimensions[r2].height = 18
            r2 += 1
            for lbl, val, unit in params:
                l = ws2.cell(r2, 2, lbl)
                l.font = lbl_font()
                l.fill = PatternFill("solid", fgColor=C_LABEL_BG)
                l.alignment = Alignment(vertical="center", indent=2)
                l.border = bdr
                disp = f"{val}  {unit}".strip() if unit else str(val)
                v = ws2.cell(r2, 3, disp)
                v.font = val_font()
                v.fill = PatternFill("solid", fgColor=C_ALT_BG)
                v.alignment = Alignment(vertical="center", indent=1)
                v.border = bdr
                ws2.row_dimensions[r2].height = 18
                r2 += 1
            r2 += 1

    wb.save(filepath)
